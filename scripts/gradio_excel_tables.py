"""Standalone Gradio demo: detect & split *stacked* tables in an Excel sheet.

Unlike ``doc2rag.ingestion.excel.load_excel_tables`` (which assumes one
fixed-schema table per sheet), this demo makes no layout assumptions. It treats
each worksheet as a raw cell grid and discovers table blocks by
**connected-components** over the occupied cells, so any number of tables
stacked **vertically** (blank row between them) or **horizontally** (blank
column between them) — in any position on the sheet — are found and separated.

For every detected block it emits:
  * sheet name, A1 cell range, inferred header row, data rows,
  * a title/caption guessed from a lone text cell sitting just above the block,
  * a Markdown rendering and a structured JSON record (ready for RAG chunking).

A colour-coded HTML "cell map" shows which cell belongs to which table — the
spreadsheet analogue of bounding-box overlays.

Run:
    python scripts/gradio_excel_tables.py
Then open the printed local URL.

Depends only on gradio + openpyxl (both already used by this project); no scipy,
pandas, or tabulate required — connected-components and Markdown are hand-rolled.
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
import tempfile
from pathlib import Path

import gradio as gr
import openpyxl
from openpyxl.utils import get_column_letter

# doc2rag lives in ./src when running from a source checkout.
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

# The connected-components table detector now lives in the package so the API /
# Modal service can reuse it (this demo is a thin UI over it). See
# doc2rag/excel/tables.py.
from doc2rag.excel.tables import (  # noqa: E402  (import after sys.path tweak)
    CheapLLM,
    FormulaEvaluator,
    SheetGrid,
    _detect_tables,
    _json_payload,
    _markdown,
    _stringify,
)


def _cell_map_html(grid: SheetGrid, tables: list[dict], text_blocks) -> str:
    """A colour-coded HTML rendering of the sheet: each table block gets its own
    background colour, non-table text blocks are grey. Shows the split at a
    glance."""
    owner = [[None] * grid.n_cols for _ in range(grid.n_rows)]
    for t in tables:
        r0, c0, r1, c1 = t["box"]
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                owner[r][c] = t["color"]
    for blk in text_blocks:
        r0, c0, r1, c1 = blk["box"]
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                if owner[r][c] is None:
                    owner[r][c] = "#9e9e9e"

    css = (
        "border-collapse:collapse;font:12px/1.3 monospace;table-layout:fixed;"
    )
    cell = "border:1px solid #ddd;padding:2px 5px;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;"
    head = "border:1px solid #ccc;padding:2px 5px;background:#f0f0f0;color:#555;text-align:center;"

    html = [f'<div style="overflow:auto;max-height:640px"><table style="{css}">']
    # column-letter header row
    html.append(f'<tr><th style="{head}"></th>')
    for c in range(grid.n_cols):
        html.append(f'<th style="{head}">{get_column_letter(c + grid.min_col)}</th>')
    html.append("</tr>")
    for r in range(grid.n_rows):
        html.append(f'<tr><th style="{head}">{r + grid.min_row}</th>')
        for c in range(grid.n_cols):
            bg = owner[r][c]
            style = cell + (f"background:{bg};color:#fff;" if bg else "background:#fff;color:#bbb;")
            txt = _stringify(grid.value[r][c])
            html.append(f'<td style="{style}" title="{txt}">{txt}</td>')
        html.append("</tr>")
    html.append("</table></div>")
    return "".join(html)



# --------------------------------------------------------------------------- engine
def parse_excel(file_path: str | None, use_summary: bool = False,
                use_recheck: bool = False):
    """Return (summary, markdown, json_str, cell_map_html, llm_log, md_file,
    json_file)."""
    empty = (None, None)
    if not file_path:
        return "Please upload an .xlsx / .xlsm file first.", "", "", "", "", *empty

    path = Path(file_path)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        # Second load keeps the formulas so we can compute cells whose cached
        # value is missing (openpyxl can't hold both views in one workbook).
        wb_f = openpyxl.load_workbook(path, data_only=False)
    except Exception as exc:  # noqa: BLE001 - surface to UI
        return f"⚠️ Failed to open workbook: `{exc}`", "", "", "", "", *empty

    use_llm = use_summary or use_recheck
    llm = CheapLLM() if use_llm else None

    all_sheets: list[dict] = []
    map_blocks: list[str] = []
    for sheet in wb.worksheets:
        form_sheet = wb_f[sheet.title] if sheet.title in wb_f.sheetnames else None
        evaluator = FormulaEvaluator(form_sheet, sheet) if form_sheet else None
        grid = SheetGrid(sheet, evaluator=evaluator)
        if grid.n_rows == 0 or grid.n_cols == 0:
            all_sheets.append({"sheet": sheet.title, "tables": []})
            continue
        tables, text_blocks = _detect_tables(
            grid, llm=llm, do_summary=use_summary, do_recheck=use_recheck
        )
        all_sheets.append({"sheet": sheet.title, "tables": tables,
                           "text_blocks": text_blocks})
        declared = ", ".join(sheet.tables.keys()) if getattr(sheet, "tables", None) else ""
        note = f' <span style="color:#888">(declared Excel Tables: {declared})</span>' if declared else ""
        map_blocks.append(
            f"<h4>Sheet: {sheet.title} — {len(tables)} table(s), "
            f"{len(text_blocks)} text block(s){note}</h4>"
            + _cell_map_html(grid, tables, text_blocks)
        )

    total = sum(len(s["tables"]) for s in all_sheets)
    llm_bit = ""
    if use_llm:
        tasks = "+".join(t for t, on in (("summary", use_summary),
                                         ("recheck", use_recheck)) if on)
        if llm and llm.ready:
            llm_bit = f" · LLM={llm.model} [{tasks}], API calls={llm.calls}"
        else:
            llm_bit = " · LLM=off (see log)"
    summary = (
        f"✅ {path.name} · sheets={len(all_sheets)} · "
        f"tables detected={total} · "
        f"method=connected-components (4-conn, merged-cell aware){llm_bit}"
    )

    # LLM activity log for the UI.
    if not use_llm:
        llm_log = (
            "_LLM off. Enable **Summary** for a one-line RAG description per "
            "table, and/or **Recheck** to have the model fix ambiguous "
            "structure (infer key:value, add/repair headers)._"
        )
    elif llm and llm.log:
        llm_log = "**LLM log**\n\n" + "\n".join(llm.log)
    elif llm and llm.ready:
        llm_log = "_LLM enabled — nothing needed rechecking / summarising._"
    else:
        llm_log = "_LLM could not start (missing key / provider)._"

    markdown = _markdown(all_sheets)
    payload = _json_payload(all_sheets)
    json_str = json.dumps(payload, ensure_ascii=False, indent=2)
    cell_map = "<hr>".join(map_blocks) or "_No sheets._"

    md_file = str(path.with_suffix(".tables.md"))
    json_file = str(path.with_suffix(".tables.json"))
    try:
        Path(md_file).write_text(markdown, encoding="utf-8")
        Path(json_file).write_text(json_str, encoding="utf-8")
    except OSError:
        md_file = json_file = None

    return summary, markdown, json_str, cell_map, llm_log, md_file, json_file


# --------------------------------------------------------------------------- UI
def build_demo() -> gr.Blocks:
    with gr.Blocks(title="Excel Stacked-Table Splitter") as demo:
        gr.Markdown(
            "# 📊 Excel Stacked-Table Splitter\n"
            "Upload an `.xlsx` / `.xlsm` file. The demo treats each sheet as a "
            "raw cell grid and finds table blocks by **connected-components**, so "
            "tables stacked **vertically or horizontally** (in any position, any "
            "count) are detected and split — no fixed schema assumed. "
            "Get a colour-coded cell map, Markdown, and RAG-ready JSON.\n\n"
            "Two independent, opt-in LLM tasks (cheap model, structure-only, "
            "cached):\n"
            "- **🤖 Summary** — a one-line RAG description per table.\n"
            "- **🔎 Recheck** — the model reviews each detected table and fixes "
            "ambiguous structure: infer a key:value list, add or repair headers, "
            "flag a header row that's really data."
        )
        with gr.Row():
            with gr.Column(scale=1):
                file_in = gr.File(
                    label="Excel workbook",
                    file_types=[".xlsx", ".xlsm"],
                    type="filepath",
                )
                use_summary = gr.Checkbox(
                    label="🤖 LLM Summary — one-line description per table",
                    value=False,
                )
                use_recheck = gr.Checkbox(
                    label="🔎 LLM Recheck — fix ambiguous table structure",
                    value=False,
                )
                run_btn = gr.Button("🔍 Detect & split tables", variant="primary")
                status = gr.Markdown("")
                md_dl = gr.File(label="⬇ Markdown (.md)")
                json_dl = gr.File(label="⬇ JSON (.json)")
            with gr.Column(scale=2):
                with gr.Tabs():
                    with gr.Tab("🗺 Cell map"):
                        map_out = gr.HTML()
                    with gr.Tab("📝 Markdown"):
                        md_out = gr.Markdown()
                    with gr.Tab("{ } JSON"):
                        json_out = gr.Code(language="json", label="Detected tables")
                    with gr.Tab("🤖 LLM log"):
                        llm_out = gr.Markdown()

        run_btn.click(
            fn=parse_excel,
            inputs=[file_in, use_summary, use_recheck],
            outputs=[status, md_out, json_out, map_out, llm_out, md_dl, json_dl],
        )
    return demo


if __name__ == "__main__":
    build_demo().launch(theme=gr.themes.Soft(), share=True)
