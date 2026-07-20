"""Headless Excel stacked-table detection.

Extracted from ``scripts/gradio_excel_tables.py`` so the same connected-components
table detector can be used by the API / Modal service without pulling in Gradio.
The Gradio demo is a thin UI over this module.

Public entry point: :func:`parse_workbook` — takes an .xlsx/.xlsm path and returns
the detected tables (RAG-ready records) plus a Markdown rendering. The optional
LLM summary/recheck layer is off by default and degrades gracefully without keys.
"""

from __future__ import annotations

import ast
import hashlib
import json
import re
import tempfile
from collections import deque
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

PALETTE = [
    "#e6194b", "#3cb44b", "#0082c8", "#f58231", "#911eb4", "#46d0d0",
    "#f032e6", "#d2f53c", "#fabebe", "#008080", "#aa6e28", "#800000",
]

# Blocks smaller than this in either dimension are treated as captions/notes
# rather than tables (a title cell, a stray footnote, etc.).
MIN_TABLE_ROWS = 2
MIN_TABLE_COLS = 2


# ------------------------------------------------------------- formula evaluation
# Files written by non-Excel tools often store formulas with NO cached result,
# so openpyxl's data_only mode returns None (blank) even though Excel would show
# a value (Margin = Revenue-COGS, a Total row of SUMs, ...). We evaluate those
# formulas ourselves: cell refs, ranges, +-*/(), and the common aggregate
# functions. Anything unsupported falls back to the raw formula text (never
# blank). No external dependency, no arbitrary-code eval (AST is whitelisted).

_TOKEN_RE = re.compile(
    r"""
    (?P<FUNC>[A-Za-z_][A-Za-z0-9_.]*)(?=\s*\() |
    (?P<RANGE>\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+) |
    (?P<CELL>\$?[A-Z]{1,3}\$?\d+) |
    (?P<NUM>\d+\.?\d*(?:[eE][-+]?\d+)?) |
    (?P<STR>"[^"]*") |
    (?P<OP>[-+*/(),]) |
    (?P<WS>\s+) |
    (?P<OTHER>.)
    """,
    re.VERBOSE,
)


def _flat_nums(args):
    out = []
    for a in args:
        if isinstance(a, (list, tuple)):
            out.extend(_flat_nums(a))
        elif isinstance(a, bool):
            continue
        elif isinstance(a, (int, float)):
            out.append(a)
    return out


_FORMULA_FUNCS = {
    "SUM": lambda *a: sum(_flat_nums(a)),
    "AVERAGE": lambda *a: (sum(_flat_nums(a)) / len(_flat_nums(a))) if _flat_nums(a) else 0,
    "MIN": lambda *a: min(_flat_nums(a)) if _flat_nums(a) else 0,
    "MAX": lambda *a: max(_flat_nums(a)) if _flat_nums(a) else 0,
    "COUNT": lambda *a: len(_flat_nums(a)),
    "ROUND": lambda x, n=0: round(x, int(n)),
    "ABS": abs,
}

_AST_OK = (
    ast.Expression, ast.BinOp, ast.UnaryOp, ast.Add, ast.Sub, ast.Mult,
    ast.Div, ast.USub, ast.UAdd, ast.Call, ast.Name, ast.Load,
    ast.Constant, ast.List, ast.Tuple,
)


def _ast_is_safe(node) -> bool:
    for n in ast.walk(node):
        if not isinstance(n, _AST_OK):
            return False
        if isinstance(n, ast.Call) and not isinstance(n.func, ast.Name):
            return False
    return True


class FormulaEvaluator:
    """Resolves a cell to its computed value: cached result if present, else the
    evaluated formula, else (on anything unsupported) the raw formula text."""

    def __init__(self, form_sheet, data_sheet=None) -> None:
        self._f = form_sheet
        self._d = data_sheet
        self._cache: dict[tuple[int, int], object] = {}
        self._active: set[tuple[int, int]] = set()

    def get(self, row: int, col: int):
        key = (row, col)
        if key in self._cache:
            return self._cache[key]
        if self._d is not None:
            cached = self._d.cell(row=row, column=col).value
            if cached is not None:
                self._cache[key] = cached
                return cached
        raw = self._f.cell(row=row, column=col).value
        if not (isinstance(raw, str) and raw.startswith("=")):
            self._cache[key] = raw
            return raw
        if key in self._active:            # circular reference guard
            return 0
        self._active.add(key)
        try:
            val = self._evaluate(raw[1:])
        except Exception:                  # noqa: BLE001 - unsupported -> show text
            val = raw
        self._active.discard(key)
        self._cache[key] = val
        return val

    def _cell(self, ref: str):
        r, c = coordinate_to_tuple(ref.replace("$", ""))
        return self.get(r, c)

    def _rng(self, ref: str):
        min_c, min_r, max_c, max_r = range_boundaries(ref.replace("$", ""))
        return [self.get(r, c)
                for r in range(min_r, max_r + 1)
                for c in range(min_c, max_c + 1)]

    def _to_python(self, expr: str) -> str:
        out = []
        for m in _TOKEN_RE.finditer(expr):
            kind, tok = m.lastgroup, m.group()
            if kind == "WS":
                out.append(" ")
            elif kind == "FUNC":
                name = tok.upper()
                if name not in _FORMULA_FUNCS:
                    raise ValueError(f"unsupported function {name}")
                out.append(name)
            elif kind == "RANGE":
                out.append(f"_rng({tok.replace('$', '')!r})")
            elif kind == "CELL":
                out.append(f"_cell({tok.replace('$', '')!r})")
            elif kind in ("NUM", "STR", "OP"):
                out.append(tok)
            else:
                raise ValueError(f"unexpected token {tok!r}")
        return "".join(out)

    def _evaluate(self, expr: str):
        py = self._to_python(expr)
        tree = ast.parse(py, mode="eval")
        if not _ast_is_safe(tree):
            raise ValueError("unsafe expression")
        ns = {"_cell": self._cell, "_rng": self._rng, **_FORMULA_FUNCS}
        return eval(compile(tree, "<formula>", "eval"), {"__builtins__": {}}, ns)

    def formula(self, row: int, col: int) -> str | None:
        """The raw formula string at a cell (e.g. '=C5-D5'), or None."""
        raw = self._f.cell(row=row, column=col).value
        return raw if isinstance(raw, str) and raw.startswith("=") else None


_AGG_FUNCS = ("SUM", "AVERAGE", "COUNT", "MIN", "MAX", "PRODUCT")


def _is_aggregate_formula(f: str) -> bool:
    """A formula that summarises many cells (a range or an aggregate call) --
    the signature of a total/average row rather than a per-row calculation."""
    u = f.upper()
    return ":" in u or any(fn + "(" in u for fn in _AGG_FUNCS)


def _formula_to_expr(grid, formula: str, headers, c0: int) -> str:
    """Rewrite a formula into a plain-language expression by replacing cell
    references with their column headers: '=C5-D5' -> 'Revenue - COGS',
    '=SUM(C5:C9)' -> 'SUM(Revenue)'."""
    def hdr(letter: str) -> str:
        gcol = column_index_from_string(letter) - grid.min_col
        hi = gcol - c0
        return headers[hi] if 0 <= hi < len(headers) else letter

    out = []
    for m in _TOKEN_RE.finditer(formula[1:] if formula.startswith("=") else formula):
        kind, tok = m.lastgroup, m.group()
        if kind == "CELL":
            out.append(hdr(re.match(r"\$?([A-Z]{1,3})", tok).group(1)))
        elif kind == "RANGE":
            a, b = tok.split(":")
            la = re.match(r"\$?([A-Z]{1,3})", a).group(1)
            lb = re.match(r"\$?([A-Z]{1,3})", b).group(1)
            out.append(hdr(la) if la == lb else tok)
        elif kind == "OP" and tok in "+-*/":
            out.append(f" {tok} ")
        elif kind == "WS":
            continue
        else:
            out.append(tok)
    return re.sub(r"\s+", " ", "".join(out)).strip()


def _extract_derived(grid, headers, c0, grid_rows, data_rows):
    """Capture formula provenance for RAG, at the RIGHT granularity:
      derived_columns -> {header: {expression, formula}} for columns computed the
                         same way in every (non-aggregate) row (Margin = Rev-COGS);
      aggregate_rows  -> [{row_index, label, cells:{header: expression}}] for
                         total/average rows that summarise other rows.
    Returns (derived_columns, aggregate_rows) -- empty if the table has no
    formulas, so plain data tables are unaffected."""
    ncols = len(headers)
    per_col: dict[int, list[tuple[int, str]]] = {i: [] for i in range(ncols)}
    agg_idx = []
    for di, gr in enumerate(grid_rows):
        val_cols = agg_cols = 0
        for i in range(ncols):
            if grid.value[gr][c0 + i] is not None:
                val_cols += 1
            f = grid.formula[gr][c0 + i]
            if not f:
                continue
            per_col[i].append((di, f))
            if _is_aggregate_formula(f):
                agg_cols += 1
        # A totals/average row is one where *most* value cells are aggregates,
        # not one that merely has a single helper formula (e.g. auto-numbering).
        if val_cols and agg_cols > 0.5 * val_cols:
            agg_idx.append(di)

    derived_columns = {}
    for i in range(ncols):
        entries = [(di, f) for di, f in per_col[i] if di not in agg_idx]
        if len(entries) < 2:
            continue
        exprs = {_formula_to_expr(grid, f, headers, c0) for _, f in entries}
        if len(exprs) == 1:
            derived_columns[headers[i]] = {
                "expression": next(iter(exprs)),
                "formula": entries[0][1],
            }

    aggregate_rows = []
    for di in agg_idx:
        gr = grid_rows[di]
        cells = {}
        for i in range(ncols):
            f = grid.formula[gr][c0 + i]
            if f:
                cells[headers[i]] = _formula_to_expr(grid, f, headers, c0)
        label = next((v for v in data_rows[di] if v), "")
        aggregate_rows.append({"row_index": di, "label": label, "cells": cells})

    return derived_columns, aggregate_rows


# --------------------------------------------------------------------- grid model
class SheetGrid:
    """A dense, 0-indexed view of a worksheet's used range.

    ``value[r][c]`` is the (data-only) cell value or ``None``. Merged cells are
    flattened so the merged value is **repeated into every covered slot** (per
    the spec: a merged cell makes all its cells share the value) and each slot
    is marked occupied so a merged header can't split a table in two.

    ``anchor[r][c]`` keeps the *pre-fill* value (present only at a merge's
    top-left), so header detection can distinguish a single merged banner
    (one anchor) from a real spanning group header (several anchors).
    ``merges`` lists each merged range as 0-indexed (r0, c0, r1, c1).
    """

    def __init__(self, sheet, evaluator=None) -> None:
        self.title = sheet.title
        self.min_row, self.min_col = sheet.min_row, sheet.min_column
        self.max_row, self.max_col = sheet.max_row, sheet.max_column
        self.n_rows = max(0, self.max_row - self.min_row + 1)
        self.n_cols = max(0, self.max_col - self.min_col + 1)

        self.value = [[None] * self.n_cols for _ in range(self.n_rows)]
        self.occupied = [[False] * self.n_cols for _ in range(self.n_rows)]
        # Non-default background fill (rgb hex) per cell, used only to infer the
        # *level* of section-divider rows -- never to identify a colour.
        self.fill = [[None] * self.n_cols for _ in range(self.n_rows)]
        # Raw formula string per cell ('=C5-D5' or None) for provenance capture.
        self.formula = [[None] * self.n_cols for _ in range(self.n_rows)]

        for row in sheet.iter_rows(
            min_row=self.min_row, max_row=self.max_row,
            min_col=self.min_col, max_col=self.max_col,
            values_only=False,
        ):
            for cell in row:
                r, c = cell.row - self.min_row, cell.column - self.min_col
                self.fill[r][c] = _fill_rgb(cell)
                v = cell.value
                # Cached value missing but the cell is a formula -> compute it,
                # so Margin/Total columns aren't blank (see FormulaEvaluator).
                if evaluator is not None:
                    self.formula[r][c] = evaluator.formula(cell.row, cell.column)
                    if v is None:
                        v = evaluator.get(cell.row, cell.column)
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                self.value[r][c] = v
                self.occupied[r][c] = True

        # Snapshot anchors (top-left-only values) before propagating merges.
        self.anchor = [row[:] for row in self.value]

        # Flatten merged ranges: repeat the top-left value into every covered
        # cell and mark the whole rectangle occupied.
        self.merges: list[tuple[int, int, int, int]] = []
        for mr in sheet.merged_cells.ranges:
            r0, c0 = mr.min_row - self.min_row, mr.min_col - self.min_col
            r1, c1 = mr.max_row - self.min_row, mr.max_col - self.min_col
            r0, c0 = max(0, r0), max(0, c0)
            r1, c1 = min(self.n_rows - 1, r1), min(self.n_cols - 1, c1)
            if r0 > r1 or c0 > c1:
                continue
            self.merges.append((r0, c0, r1, c1))
            merged_val = self.value[r0][c0]
            for r in range(r0, r1 + 1):
                for c in range(c0, c1 + 1):
                    self.occupied[r][c] = True
                    if merged_val is not None:
                        self.value[r][c] = merged_val

    def a1(self, r0: int, c0: int, r1: int, c1: int) -> str:
        """0-indexed block bounds -> A1 range, e.g. 'B2:E10'."""
        top = f"{get_column_letter(c0 + self.min_col)}{r0 + self.min_row}"
        bot = f"{get_column_letter(c1 + self.min_col)}{r1 + self.min_row}"
        return f"{top}:{bot}"


def _fill_rgb(cell) -> str | None:
    """Return a cell's solid background colour as an 'RRGGBB'-ish hex, or None
    for no-fill / white / theme-only fills we can't resolve to a stable hex."""
    fill = getattr(cell, "fill", None)
    if not fill or fill.patternType is None:
        return None
    rgb = getattr(fill.fgColor, "rgb", None)
    if isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
        return rgb
    return None


# --------------------------------------------------- section-hierarchy detection
# Within a table, a row carrying exactly ONE distinct value (anchor) is a
# section-divider banner, not data. Its nesting *level* is inferred generically
# from a style stack -- a full-width merged banner is a top-level section; other
# single-value rows nest by fill colour via containment. No colour is hardcoded,
# so the same logic works for any convention (green>orange here, anything else
# elsewhere).

def _anchor_count(grid, r, c0, c1) -> int:
    return sum(1 for c in range(c0, c1 + 1) if grid.anchor[r][c] is not None)


def _full_width_merge(grid, r, c0, c1) -> bool:
    """True if a single merge covers most of the row within [c0, c1]."""
    width = c1 - c0 + 1
    for mr0, mc0, mr1, mc1 in grid.merges:
        if mr0 <= r <= mr1 and (min(mc1, c1) - max(mc0, c0) + 1) >= 0.6 * width \
                and mc0 <= c0 + 1:
            return True
    return False


def _is_title_banner(grid, r, c0, c1) -> bool:
    """True if row r is a title banner: a wide horizontal merge near the block's
    left edge whose columns are NOT densely continued by the row below. That
    last check is what separates a title (row below starts unrelated content)
    from a 2-row group header like '2025' over 'Q1 Q2 Q3 Q4' (row below fills
    the merged span), which must be kept."""
    width = c1 - c0 + 1
    for mr0, mc0, mr1, mc1 in grid.merges:
        if not (mr0 <= r <= mr1 and mc1 > mc0):        # horizontal merge on row r
            continue
        span = min(mc1, c1) - max(mc0, c0) + 1
        if span < 0.5 * width or mc0 > c0 + 1:
            continue
        if r + 1 >= grid.n_rows:
            return True
        below = sum(
            1 for c in range(max(mc0, c0), min(mc1, c1) + 1)
            if grid.value[r + 1][c] is not None
        )
        if below < 0.5 * (min(mc1, c1) - max(mc0, c0) + 1):
            return True
    return False


def _divider_style(grid, r, c0, c1):
    """A hashable style for a divider row: (is_full_width_merge, fill_rgb)."""
    fw = _full_width_merge(grid, r, c0, c1)
    rgb = None
    for c in range(c0, c1 + 1):
        if grid.anchor[r][c] is not None:
            rgb = grid.fill[r][c]
            break
    return (fw, rgb)


def _divider_title(grid, r, c0, c1) -> str:
    for c in range(c0, c1 + 1):
        if grid.anchor[r][c] is not None:
            return _stringify(grid.anchor[r][c])
    return ""


def _is_strong_header(grid, r, c0, c1) -> bool:
    """A row that looks like column headers: several filled cells, all text."""
    filled = [c for c in range(c0, c1 + 1) if grid.value[r][c] is not None]
    if len(filled) < max(3, 0.5 * (c1 - c0 + 1)):
        return False
    return all(not _is_number(grid.value[r][c]) for c in filled)


def _build_sections(grid, r0, c0, r1, c1, headers):
    """Walk data rows r0..r1, splitting them into a nested section tree at
    divider rows. Returns (sections_tree, flat_records)."""
    tree: list[dict] = []
    node_stack: list[dict] = []      # currently-open section nodes (by depth)
    style_stack: list = []           # divider styles (for level inference)
    flat: list[dict] = []

    def close_to(level, end_row):
        while node_stack and node_stack[-1]["level"] >= level:
            node_stack.pop()["end"] = end_row

    def level_for(style, is_merge):
        if is_merge:                       # full-width banner => top-level
            style_stack.clear()
            style_stack.append(style)
            return 1
        if style in style_stack:           # seen before => its established level
            i = style_stack.index(style)
            del style_stack[i + 1:]
            return i + 1
        style_stack.append(style)          # new style => one level deeper
        return len(style_stack)

    for r in range(r0, r1 + 1):
        ac = _anchor_count(grid, r, c0, c1)
        if ac == 0:
            continue                       # blank spacer row
        if ac == 1:                        # divider row -> open a section
            fw = _full_width_merge(grid, r, c0, c1)
            level = level_for(_divider_style(grid, r, c0, c1), fw)
            close_to(level, r - 1)
            node = {
                "title": _divider_title(grid, r, c0, c1),
                "level": level,
                "start": r,
                "end": r1,
                "records": [],
                "subsections": [],
            }
            (node_stack[-1]["subsections"] if node_stack else tree).append(node)
            node_stack.append(node)
            continue
        # data row -> record under the deepest open section (or table root)
        cells = [_stringify(grid.value[r][c]) for c in range(c0, c1 + 1)]
        rec = dict(zip(headers, cells))
        flat.append(rec)
        if node_stack:
            node_stack[-1]["records"].append(rec)

    # Attach A1 ranges to every section now that ends are known.
    def stamp(node):
        node["range"] = grid.a1(node["start"], c0, node["end"], c1)
        node.pop("start", None)
        node.pop("end", None)
        for s in node["subsections"]:
            stamp(s)
    for n in tree:
        stamp(n)
    return tree, flat


def _has_dividers(grid, r0, c0, r1, c1) -> bool:
    """Worth building a section tree? Need a wide block with either a
    full-width merged divider or >=2 single-value divider rows."""
    if c1 - c0 + 1 < 4:
        return False
    singles = fw = 0
    for r in range(r0, r1 + 1):
        if _anchor_count(grid, r, c0, c1) == 1:
            singles += 1
            if _full_width_merge(grid, r, c0, c1):
                fw += 1
    return fw >= 1 or singles >= 2


def _has_horizontal_group(grid, r, c0, c1) -> bool:
    """True if row r has a merge spanning >1 column within [c0, c1] — a group
    header sitting above per-column sub-headers."""
    for mr0, mc0, mr1, mc1 in grid.merges:
        if mr0 <= r <= mr1 and mc1 > mc0 and mc0 <= c1 and mc1 >= c0:
            return True
    return False


# ---------------------------------------------------- header-anchored detection
# Sparse, wide tables (a body broken into sections by divider banners) can't be
# found by cell adjacency -- connected-components fragments them, and header
# detection by "density" is unreliable (a fully-filled data row looks like a
# header). Instead we anchor on the DIVIDER banners, which are unambiguous
# (a full-width merge, or a single-value coloured row), then take the header as
# the dense text row directly above the first divider. Robust to whether data
# rows are sparse or dense.

def _is_sheet_divider(grid, r) -> bool:
    """A section-divider banner spanning the whole sheet width: exactly one
    value in the row, and either a full-width merge or a coloured fill."""
    c0, c1 = 0, grid.n_cols - 1
    if _anchor_count(grid, r, c0, c1) != 1:
        return False
    if _full_width_merge(grid, r, c0, c1):
        return True
    for c in range(c0, c1 + 1):
        if grid.anchor[r][c] is not None:
            return grid.fill[r][c] is not None
    return False


def _dense_text_row(grid, r, c0, c1) -> bool:
    """A plausible header row: >=3 filled cells, all text, covering most of the
    span. (Used only to validate the row sitting above the first divider.)"""
    filled = [c for c in range(c0, c1 + 1) if grid.value[r][c] is not None]
    if len(filled) < 3 or len(filled) < 0.5 * (c1 - c0 + 1):
        return False
    return not any(_is_number(grid.value[r][c]) for c in filled)


def _find_structured_regions(grid):
    """Return (r0, c0, r1, c1) for each divider-anchored, section-divided table.

    Anchored on divider banners, header = dense row just above the first one."""
    W = grid.n_cols
    divs = [r for r in range(grid.n_rows) if _is_sheet_divider(grid, r)]
    regions = []
    used = set()
    for d in divs:
        if d in used:
            continue
        # Header = nearest dense text row above this divider (skipping blanks).
        hr = d - 1
        while hr >= 0 and _anchor_count(grid, hr, 0, W - 1) == 0:
            hr -= 1
        if hr < 0 or not _dense_text_row(grid, hr, 0, W - 1):
            used.add(d)
            continue
        acols = [c for c in range(W) if grid.value[hr][c] is not None]
        c0, c1 = min(acols), max(acols)
        # Extend down while rows keep carrying content in the span (stop at a
        # fully-blank row / end): that captures every section of this table.
        end = d
        rr = d
        while rr < grid.n_rows and _anchor_count(grid, rr, c0, c1) != 0:
            end = rr
            rr += 1
        for x in divs:                    # all dividers inside this block are done
            if hr <= x <= end:
                used.add(x)
        regions.append((hr, c0, end, c1))
    return regions


# ------------------------------------------------------------ connected components
def _components(occupied: list[list[bool]], n_rows: int, n_cols: int):
    """4-connectivity flood fill over occupied cells.

    Returns (labels, blocks) where ``labels[r][c]`` is a block id (or -1) and
    ``blocks`` maps id -> bounding box (r0, c0, r1, c1). Any fully-blank row or
    column between two regions of cells breaks connectivity, which is exactly
    how tables stacked vertically or horizontally get separated.
    """
    labels = [[-1] * n_cols for _ in range(n_rows)]
    blocks: dict[int, list[int]] = {}
    nxt = 0
    for sr in range(n_rows):
        for sc in range(n_cols):
            if not occupied[sr][sc] or labels[sr][sc] != -1:
                continue
            q = deque([(sr, sc)])
            labels[sr][sc] = nxt
            r0 = r1 = sr
            c0 = c1 = sc
            while q:
                r, c = q.popleft()
                r0, r1 = min(r0, r), max(r1, r)
                c0, c1 = min(c0, c), max(c1, c)
                for dr, dc in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                    nr, nc = r + dr, c + dc
                    if (
                        0 <= nr < n_rows and 0 <= nc < n_cols
                        and occupied[nr][nc] and labels[nr][nc] == -1
                    ):
                        labels[nr][nc] = nxt
                        q.append((nr, nc))
            blocks[nxt] = [r0, c0, r1, c1]
            nxt += 1
    return labels, blocks


# --------------------------------------------- split touching (adjacent) tables
# Connected-components can't separate tables that touch with no blank row/column.
# These value-based refiners split the two tractable cases:
#   * vertical  — a header-like row reappearing *after* data starts a new table;
#   * horizontal — a header row whose labels repeat with a fixed period
#                  (Dept,Count,Dept,Count) is two side-by-side tables.
# Horizontally adjacent tables with *different* headers and no gap can't be
# split from values alone — that needs formatting (borders); see notes.

def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _split_vertical(grid, r0, c0, r1, c1):
    """Split a block into stacked sub-tables at internal header rows."""
    header = [_stringify(grid.value[r0][c]) for c in range(c0, c1 + 1)]
    boundaries = [r0]
    seen_data = False
    for r in range(r0 + 1, r1 + 1):
        filled = [c for c in range(c0, c1 + 1) if grid.value[r][c] is not None]
        if not filled:
            continue
        numeric = [c for c in filled if _is_number(grid.value[r][c])]
        row_text = [_stringify(grid.value[r][c]) for c in range(c0, c1 + 1)]
        header_like = len(filled) >= 2 and not numeric      # text-only, multi-cell
        repeats_header = row_text == header and any(header)
        if seen_data and (repeats_header or header_like):
            boundaries.append(r)          # data -> header transition: new table
            seen_data = False
        elif numeric or not header_like:
            seen_data = True              # a data row within the current table
    if len(boundaries) == 1:
        return None
    regions = []
    for i, b in enumerate(boundaries):
        end = boundaries[i + 1] - 1 if i + 1 < len(boundaries) else r1
        regions.append((b, c0, end, c1))
    return regions


def _split_horizontal(grid, r0, c0, r1, c1):
    """Split a block into side-by-side sub-tables when the header row repeats
    with a fixed column period (identical headers)."""
    header = [_stringify(grid.value[r0][c]) for c in range(c0, c1 + 1)]
    width = len(header)
    for p in range(1, width // 2 + 1):
        if width % p:
            continue
        repeats = all(header[i] == header[i % p] for i in range(width))
        if repeats and any(header[:p]):
            return [(r0, c0 + k, r1, c0 + k + p - 1) for k in range(0, width, p)]
    return None


def _refine_block(grid, region, depth=0):
    """Recursively split one connected-component block into real tables."""
    r0, c0, r1, c1 = region
    if depth > 8 or r1 - r0 < MIN_TABLE_ROWS:
        return [region]
    for splitter in (_split_horizontal, _split_vertical):
        parts = splitter(grid, r0, c0, r1, c1)
        if parts and len(parts) > 1:
            out = []
            for p in parts:
                out.extend(_refine_block(grid, p, depth + 1))
            return out
    return [region]


# ----------------------------------------------------------------- table building
def _stringify(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _uniquify(headers: list[str]) -> list[str]:
    """Suffix duplicate headers (h, h_2, h_3, ...) so record keys stay distinct."""
    seen: dict[str, int] = {}
    out = []
    for h in headers:
        n = seen.get(h, 0) + 1
        seen[h] = n
        out.append(h if n == 1 else f"{h}_{n}")
    return out


def _build_table(grid, r0, c0, r1, c1):
    """Build one region into a table dict, or classify it as a caption.

    Returns ('table', dict) | ('caption', (box, text)) | None (empty)."""
    # Peel leading title rows: a row holding a single merged/standalone value
    # (one anchor) is a caption, not part of the header. A row with several
    # anchors (e.g. "2025 | 2026") is a real spanning header and is kept.
    inline_title_parts: list[str] = []
    while r1 - r0 + 1 > MIN_TABLE_ROWS and (
        _anchor_count(grid, r0, c0, c1) <= 1
        or _is_title_banner(grid, r0, c0, c1)
    ):
        for c in range(c0, c1 + 1):
            if grid.anchor[r0][c] is not None:
                inline_title_parts.append(_stringify(grid.anchor[r0][c]))
        r0 += 1

    # Trim blank leading/trailing rows and columns left behind after peeling
    # (e.g. a title banner that spanned columns the data below doesn't use).
    used_c = [c for c in range(c0, c1 + 1)
              if any(grid.value[r][c] is not None for r in range(r0, r1 + 1))]
    used_r = [r for r in range(r0, r1 + 1)
              if any(grid.value[r][c] is not None for c in range(c0, c1 + 1))]
    if used_c and used_r:
        c0, c1 = min(used_c), max(used_c)
        r0, r1 = min(used_r), max(used_r)

    h, w = r1 - r0 + 1, c1 - c0 + 1
    if h < MIN_TABLE_ROWS or w < MIN_TABLE_COLS:
        text = " ".join(
            _stringify(grid.value[r][c])
            for r in range(r0, r1 + 1)
            for c in range(c0, c1 + 1)
            if grid.value[r][c] is not None
        ).strip()
        return ("caption", ([r0, c0, r1, c1], text)) if text else None

    # Header depth: a top row carrying a horizontal group merge means a two-row
    # header (group + sub-header); otherwise a single row.
    hdr_n = 2 if (h > MIN_TABLE_ROWS and _has_horizontal_group(grid, r0, c0, c1)) else 1

    # Combined header per column: join header rows top-to-bottom, collapsing
    # consecutive repeats (group "2025" over Q1..Q4 -> "2025 Q1", ...).
    headers = []
    for c in range(c0, c1 + 1):
        parts: list[str] = []
        for r in range(r0, r0 + hdr_n):
            v = _stringify(grid.value[r][c])
            if v and (not parts or parts[-1] != v):
                parts.append(v)
        headers.append(" ".join(parts))
    headers = _uniquify([hd or f"col_{i + 1}" for i, hd in enumerate(headers)])

    data_rows = []
    records = []
    grid_rows = []
    for r in range(r0 + hdr_n, r1 + 1):
        cells = [_stringify(grid.value[r][c]) for c in range(c0, c1 + 1)]
        if not any(cells):
            continue
        data_rows.append(cells)
        records.append(dict(zip(headers, cells)))
        grid_rows.append(r)

    # If the body is broken up by divider banners, recover the nested section
    # tree (level 1 = full-width merged rows, deeper = single-value coloured
    # rows). Otherwise it's a flat table.
    sections = None
    if _has_dividers(grid, r0 + hdr_n, c0, r1, c1):
        sections, flat = _build_sections(grid, r0 + hdr_n, c0, r1, c1, headers)
        if flat:
            records = flat

    # Formula provenance (RAG-ready): column-level derivations + aggregate rows.
    derived_columns, aggregate_rows = _extract_derived(
        grid, headers, c0, grid_rows, data_rows
    )

    header_cells = [_stringify(grid.value[r0][c]) for c in range(c0, c1 + 1)]
    return ("table", {
        "box": [r0, c0, r1, c1],
        "range": grid.a1(r0, c0, r1, c1),
        "headers": headers,
        "header_cells": header_cells,
        "data_rows": data_rows,
        "records": records,
        "sections": sections,
        "derived_columns": derived_columns,
        "aggregate_rows": aggregate_rows,
        "inline_title": " ".join(inline_title_parts) or None,
    })


def _recheck_candidate(table: dict) -> bool:
    """Which tables are worth an LLM recheck (keeps cost low). Section-structured
    tables are already well-formed; we target flat tables that look ambiguous:
    a header cell that is numeric (so it's probably data), or a narrow block."""
    if table.get("sections"):
        return False
    headers = table["headers"]
    numeric_hdr = any(_is_number_str(h) for h in headers)
    return numeric_hdr or len(headers) <= 3


def _is_number_str(s: str) -> bool:
    try:
        float(str(s).replace(",", ""))
        return True
    except (TypeError, ValueError):
        return False


def _apply_recheck(table: dict, res: dict) -> None:
    """Re-shape a table in place from the LLM's structural verdict. Only cells
    already extracted are rearranged -- values are never taken from the model."""
    ncols = len(table["headers"])
    first_is_data = bool(res.get("first_row_is_data"))
    all_rows = ([table["header_cells"]] if first_is_data else []) + table["data_rows"]

    if res.get("structure") == "key_value":
        ki = res.get("key_index")
        vi = res.get("value_index")
        if not isinstance(ki, int) or not isinstance(vi, int):
            ki, vi = 0, 1
        if not (0 <= ki < ncols and 0 <= vi < ncols):
            return
        recs = [
            {"key": row[ki], "value": row[vi]}
            for row in all_rows if row[ki] or row[vi]
        ]
        if not recs:
            return
        table["structure"] = "key_value"
        table["headers"] = ["key", "value"]
        table["data_rows"] = [[r["key"], r["value"]] for r in recs]
        table["records"] = recs
    else:
        new_headers = res.get("headers")
        if not (isinstance(new_headers, list) and len(new_headers) == ncols):
            new_headers = table["headers"]
        new_headers = _uniquify([_stringify(h) or f"col_{i + 1}"
                                 for i, h in enumerate(new_headers)])
        if not first_is_data and new_headers == table["headers"]:
            table["recheck_note"] = res.get("explanation")
            return
        table["structure"] = "table"
        table["headers"] = new_headers
        table["data_rows"] = all_rows
        table["records"] = [dict(zip(new_headers, row)) for row in all_rows]
    table["recheck_note"] = res.get("explanation")


def _detect_tables(grid: SheetGrid, llm=None, do_summary=False, do_recheck=False):
    """Split one sheet into table blocks + caption blocks.

    Returns (tables, captions):
      tables   -> list of dicts with range/headers/rows/records/box/color
      captions -> list of (box, text) small text-only blocks used as titles
    """
    tables: list[dict] = []
    captions: list[tuple[list[int], str]] = []
    consumed = [[False] * grid.n_cols for _ in range(grid.n_rows)]

    # 1) Header-anchored, section-divided tables first (adjacency-agnostic).
    for r0, c0, r1, c1 in _find_structured_regions(grid):
        kind = _build_table(grid, r0, c0, r1, c1)
        if kind and kind[0] == "table":
            tables.append(kind[1])
            for r in range(r0, r1 + 1):
                for c in range(c0, c1 + 1):
                    consumed[r][c] = True

    # 2) Everything else via connected-components over the UNCLAIMED cells.
    occ = [
        [grid.occupied[r][c] and not consumed[r][c] for c in range(grid.n_cols)]
        for r in range(grid.n_rows)
    ]
    _, blocks = _components(occ, grid.n_rows, grid.n_cols)
    regions = []
    for block in blocks.values():
        regions.extend(_refine_block(grid, tuple(block)))

    for r0, c0, r1, c1 in regions:
        kind = _build_table(grid, r0, c0, r1, c1)
        if not kind:
            continue
        (tables if kind[0] == "table" else captions).append(kind[1])

    # Attach a title: prefer a peeled inline banner, else a caption block sitting
    # directly above the table with overlapping columns.
    for t in tables:
        if t.get("inline_title"):
            t["title"] = t["inline_title"]
            continue
        tr0, tc0, _, tc1 = t["box"]
        best = None
        for (cr0, cc0, cr1, cc1), text in captions:
            if cr1 < tr0 and not (cc1 < tc0 or cc0 > tc1):  # above & overlapping cols
                if best is None or cr1 > best[0]:
                    best = (cr1, text)
        t["title"] = best[1] if best else None

    # Order tables top-to-bottom, then left-to-right (natural reading order).
    tables.sort(key=lambda t: (t["box"][0], t["box"][1]))
    for i, t in enumerate(tables):
        t["color"] = PALETTE[i % len(PALETTE)]

    have_llm = llm is not None and llm.ready

    # LLM task 1 (recheck): fix structure of ambiguous tables before summarising.
    if have_llm and do_recheck:
        for t in tables:
            if _recheck_candidate(t):
                res = llm.recheck(t)
                if res:
                    _apply_recheck(t, res)

    # LLM task 2 (summary): one cached, structure-only description per table.
    for t in tables:
        t["description"] = llm.describe(t) if (have_llm and do_summary) else None

    return tables, captions


# ------------------------------------------------------------------------ renderers
def _md_table(headers: list[str], rows: list[list[str]]) -> str:
    def esc(s: str) -> str:
        return s.replace("|", "\\|").replace("\n", " ")

    out = ["| " + " | ".join(esc(h) for h in headers) + " |"]
    out.append("|" + "|".join("---" for _ in headers) + "|")
    for row in rows:
        out.append("| " + " | ".join(esc(c) for c in row) + " |")
    return "\n".join(out)


def _markdown(all_sheets: list[dict]) -> str:
    md: list[str] = ["# Excel — detected tables", ""]
    total = sum(len(s["tables"]) for s in all_sheets)
    md.append(f"Detected **{total}** table(s) across **{len(all_sheets)}** sheet(s).\n")
    for s in all_sheets:
        md.append(f"## Sheet: `{s['sheet']}` — {len(s['tables'])} table(s)\n")
        if not s["tables"]:
            md.append("_No table-shaped blocks found._\n")
        for i, t in enumerate(s["tables"]):
            title = f" — {t['title']}" if t.get("title") else ""
            kv = " · key:value" if t.get("structure") == "key_value" else ""
            md.append(f"### Table {i + 1} `[{t['range']}]`{title}{kv}\n")
            if t.get("description"):
                md.append(f"> 🤖 {t['description']}\n")
            if t.get("recheck_note"):
                md.append(f"> 🔎 {t['recheck_note']}\n")
            for col, d in (t.get("derived_columns") or {}).items():
                md.append(f"> 🧮 **{col}** = {d['expression']}  (`{d['formula']}`)\n")
            for a in (t.get("aggregate_rows") or []):
                md.append(f"> Σ aggregate row **{a['label']}** (summarises other rows)\n")
            if t.get("sections"):
                _md_sections(md, t["sections"], t["headers"], depth=0)
            else:
                md.append(_md_table(t["headers"], t["data_rows"]))
            md.append("")
    return "\n".join(md)


def _md_sections(md: list[str], sections: list[dict], headers, depth: int) -> None:
    """Render a nested section tree: '####'+ heading per section, then its own
    rows as a table, then its subsections (recursively deeper)."""
    for sec in sections:
        hashes = "#" * min(6, 4 + depth)
        md.append(f"{hashes} {sec['title']} `[{sec['range']}]`\n")
        if sec["records"]:
            rows = [[rec.get(h, "") for h in headers] for rec in sec["records"]]
            md.append(_md_table(headers, rows))
            md.append("")
        if sec["subsections"]:
            _md_sections(md, sec["subsections"], headers, depth + 1)


def _json_payload(all_sheets: list[dict]) -> list[dict]:
    payload = []
    for s in all_sheets:
        for i, t in enumerate(s["tables"]):
            entry = {
                "sheet": s["sheet"],
                "table_index": i,
                "range": t["range"],
                "title": t.get("title"),
                "structure": t.get("structure", "table"),
                "description": t.get("description"),
                "recheck_note": t.get("recheck_note"),
                "n_rows": len(t["data_rows"]),
                "n_cols": len(t["headers"]),
                "headers": t["headers"],
            }
            if t.get("derived_columns"):
                entry["derived_columns"] = t["derived_columns"]
            if t.get("aggregate_rows"):
                entry["aggregate_rows"] = t["aggregate_rows"]
            if t.get("sections"):
                entry["sections"] = t["sections"]
            else:
                entry["records"] = t["records"]
            payload.append(entry)
    return payload

_CACHE_PATH = Path(tempfile.gettempdir()) / "doc2rag_excel_llm_cache.json"


def _load_cache() -> dict:
    try:
        return json.loads(_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}


def _save_cache(cache: dict) -> None:
    try:
        _CACHE_PATH.write_text(json.dumps(cache), encoding="utf-8")
    except OSError:
        pass


def _parse_json(text: str | None):
    if not text:
        return None
    try:
        return json.loads(text)
    except ValueError:
        # Tolerate ```json fences / prose around the object.
        i, j = text.find("{"), text.rfind("}")
        if 0 <= i < j:
            try:
                return json.loads(text[i:j + 1])
            except ValueError:
                return None
    return None


class CheapLLM:
    """Thin provider-agnostic JSON client that reuses the project's settings but
    defaults to the cheap model tier. Disables itself gracefully (no key / no
    deps / import error) and records every action in ``log``."""

    OPENAI_MODEL = "gpt-4o-mini"
    GEMINI_MODEL = "gemini-1.5-flash"

    def __init__(self) -> None:
        self.log: list[str] = []
        self.calls = 0
        self._provider = None
        self._client = None
        self.model = None
        self._cache = _load_cache()
        try:
            from doc2rag.config import get_settings

            s = get_settings()
        except Exception as exc:  # noqa: BLE001
            self.log.append(f"⚠️ LLM off — settings unavailable: `{exc}`")
            return
        self._provider = s.llm_provider
        try:
            if s.llm_provider == "openai":
                if not s.openai_api_key:
                    raise RuntimeError("no OpenAI API key configured")
                from openai import OpenAI

                self._client = OpenAI(api_key=s.openai_api_key)
                self.model = self.OPENAI_MODEL
            elif s.llm_provider == "gemini":
                if not s.gemini_api_key:
                    raise RuntimeError("no Gemini API key configured")
                from google import genai

                self._client = genai.Client(api_key=s.gemini_api_key)
                self.model = self.GEMINI_MODEL
            else:
                raise RuntimeError(f"unknown provider {s.llm_provider!r}")
        except Exception as exc:  # noqa: BLE001
            self.log.append(f"⚠️ LLM off — {exc}")
            self._client = None

    @property
    def ready(self) -> bool:
        return self._client is not None

    def _complete(self, system: str, user: str, kind: str):
        key = hashlib.sha1(
            f"{kind}|{self.model}|{system}|{user}".encode()
        ).hexdigest()
        if key in self._cache:
            self.log.append(f"• {kind}: cache hit (no API call)")
            return self._cache[key]
        self.calls += 1
        try:
            if self._provider == "openai":
                r = self._client.chat.completions.create(
                    model=self.model,
                    temperature=0,
                    response_format={"type": "json_object"},
                    messages=[
                        {"role": "system", "content": system},
                        {"role": "user", "content": user},
                    ],
                )
                text = r.choices[0].message.content
            else:
                from google.genai import types

                r = self._client.models.generate_content(
                    model=self.model,
                    contents=f"{system}\n\n{user}",
                    config=types.GenerateContentConfig(
                        response_mime_type="application/json"
                    ),
                )
                text = r.text
        except Exception as exc:  # noqa: BLE001
            self.log.append(f"• {kind}: API call failed — `{exc}`")
            return None
        data = _parse_json(text)
        if data is not None:
            self._cache[key] = data
            _save_cache(self._cache)
        return data

    # -- task 1: recheck a detected table's structure --------------------------
    def recheck(self, table: dict) -> dict | None:
        """Review one detected table and return structure corrections. The model
        only classifies + names; it never re-emits cell values. Returns a dict:
          structure          : "table" | "key_value"
          first_row_is_data  : bool  (the detected header row is really data)
          headers            : [str] improved/added column names (table mode)
          key_index/value_index : ints (key_value mode, 0-based columns)
          explanation        : one line describing the fix
        Sends headers + up to 12 rows only."""
        rows = table["data_rows"][:12]
        ncols = len(table["headers"])
        user = (
            "Review this table extracted from a spreadsheet and correct its "
            "structure so it is easy to understand. Decide:\n"
            "- Is it a normal table, or a KEY:VALUE list (label in one column, "
            "value in another)?\n"
            "- Is the detected header row actually a data row?\n"
            "- Are the column headers missing/wrong? Propose good ones.\n"
            f"detected_headers ({ncols} cols): {table['headers']}\n"
            f"rows: {rows}\n\n"
            'Return JSON: {"structure":"table"|"key_value",'
            '"first_row_is_data":bool,'
            f'"headers":[{ncols} strings],'
            '"key_index":int|null,"value_index":int|null,'
            '"explanation":"one short line"}'
        )
        data = self._complete(
            "You audit and normalise spreadsheet table structure. "
            "Answer only with the requested JSON.",
            user,
            "recheck",
        )
        if not isinstance(data, dict) or "structure" not in data:
            return None
        note = str(data.get("explanation") or data["structure"])
        self.log.append(f"• recheck [{table['range']}]: {note[:70]}")
        return data

    # -- task 2: one-line RAG description --------------------------------------
    def describe(self, table: dict) -> str | None:
        preview = table["records"][:2]
        user = (
            "Write ONE concise sentence describing what this table contains, "
            "for use as retrieval context in a RAG system. Do not list every "
            "column; summarize.\n"
            f"title: {table.get('title')!r}\n"
            f"headers: {table['headers']}\n"
            f"sample_rows: {preview}\n"
            'Return JSON {"description": "..."}.'
        )
        data = self._complete(
            "You write terse, factual one-sentence table summaries. "
            "Answer only with the requested JSON.",
            user,
            "describe",
        )
        try:
            desc = str(data["description"]).strip()
        except (TypeError, KeyError):
            return None
        if desc:
            snippet = desc if len(desc) <= 60 else desc[:57] + "…"
            self.log.append(f"• describe [{table['range']}]: {snippet}")
        return desc or None


def parse_workbook(path, use_summary: bool = False, use_recheck: bool = False) -> dict:
    """Detect stacked tables in every worksheet of an .xlsx/.xlsm file.

    Args:
        path: Path to the workbook.
        use_summary: If True, add a one-line LLM description per table (opt-in).
        use_recheck: If True, let the LLM fix ambiguous table structure (opt-in).

    Returns:
        ``{"sheets": [...], "num_tables": int, "tables": [...], "markdown": str}``
        where ``tables`` is the RAG-ready JSON payload (one entry per detected
        table, with headers + records / sections). Deterministic by default; the
        LLM layer only runs when a toggle is on and a provider key is available.
    """
    path = Path(path)
    workbook = openpyxl.load_workbook(path, data_only=True)
    # Second load keeps formulas so we can compute cells whose cached value is
    # missing (openpyxl can't hold both views in one workbook).
    formula_workbook = openpyxl.load_workbook(path, data_only=False)

    llm = CheapLLM() if (use_summary or use_recheck) else None
    all_sheets: list[dict] = []
    for sheet in workbook.worksheets:
        form_sheet = formula_workbook[sheet.title] if sheet.title in formula_workbook.sheetnames else None
        evaluator = FormulaEvaluator(form_sheet, sheet) if form_sheet else None
        grid = SheetGrid(sheet, evaluator=evaluator)
        if grid.n_rows == 0 or grid.n_cols == 0:
            all_sheets.append({"sheet": sheet.title, "tables": []})
            continue
        tables, _captions = _detect_tables(
            grid, llm=llm, do_summary=use_summary, do_recheck=use_recheck
        )
        all_sheets.append({"sheet": sheet.title, "tables": tables})

    return {
        "sheets": [s["sheet"] for s in all_sheets],
        "num_tables": sum(len(s["tables"]) for s in all_sheets),
        "tables": _json_payload(all_sheets),
        "markdown": _markdown(all_sheets),
    }
